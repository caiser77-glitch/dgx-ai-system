#!/usr/bin/env python3
"""Extract text, table outputs, and metadata from uploaded local files."""
import argparse
import os
import subprocess
import csv
import hashlib
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path
import re
import sys
import urllib.error
import urllib.request
import uuid


TEXT_EXTENSIONS = {".txt", ".md"}
CSV_EXTENSIONS = {".csv"}
XLSX_EXTENSIONS = {".xlsx"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
PDF_EXTENSIONS = {".pdf"}

TAXON_GROUPS = [
    "양서파충류",
    "조류",
    "포유류",
    "어류",
    "식물상",
    "육상곤충",
    "저서성대형무척추동물",
    "담수조류",
    "해충",
]

DOCUMENT_TYPE_KEYWORDS = [
    "동식물상",
    "식물상",
    "부록",
    "사업의 개요",
    "사업개요",
    "검토의견",
    "문헌",
    "현장사진",
    "사진",
    "야장",
    "영수증",
    "출장복명서",
    "보고서",
    "조사표",
]


def now_iso() -> str:
    kst = timezone(timedelta(hours=9))
    return datetime.now(kst).isoformat(timespec="seconds")


def safe_stem(source_path: Path) -> str:
    digest = hashlib.sha1(str(source_path).encode("utf-8")).hexdigest()[:10]
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", source_path.stem).strip("._")
    return f"{cleaned or 'file'}_{digest}"


def ensure_dirs(output_dir: Path) -> dict[str, Path]:
    paths = {
        "text": output_dir / "text",
        "tables": output_dir / "tables",
        "metadata": output_dir / "metadata",
    }
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def write_text(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def read_text_file(source_path: Path) -> str:
    try:
        return source_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return source_path.read_text(encoding="utf-8", errors="replace")


def extract_text(source_path: Path, dirs: dict[str, Path], stem: str) -> dict:
    output_path = dirs["text"] / f"{stem}.txt"
    write_text(output_path, read_text_file(source_path))
    return {"text": str(output_path), "tables": []}


def read_csv_rows_with_fallback(source_path: Path) -> list:
    """국내 구버전 문헌 CSV는 UTF-8이 아닌 CP949/EUC-KR로 저장된 경우가 많아
    순서대로 시도하고, 전부 실패하면 손상 바이트를 치환해서라도 읽어낸다."""
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            with source_path.open("r", encoding=encoding, newline="") as src:
                return list(csv.reader(src))
        except UnicodeDecodeError:
            continue
    with source_path.open("r", encoding="utf-8", errors="replace", newline="") as src:
        return list(csv.reader(src))


def extract_csv(source_path: Path, dirs: dict[str, Path], stem: str) -> dict:
    table_path = dirs["tables"] / f"{stem}.csv"
    rows = read_csv_rows_with_fallback(source_path)

    with table_path.open("w", encoding="utf-8", newline="") as dst:
        writer = csv.writer(dst)
        writer.writerows(rows)

    text_path = dirs["text"] / f"{stem}.txt"
    row_count = max(len(rows) - 1, 0)
    headers = rows[0] if rows else []
    summary = [
        f"source_name: {source_path.name}",
        f"source_type: csv",
        f"row_count: {row_count}",
        f"columns: {', '.join(headers)}",
    ]
    write_text(text_path, "\n".join(summary) + "\n")
    return {"text": str(text_path), "tables": [str(table_path)]}


def extract_xlsx(source_path: Path, dirs: dict[str, Path], stem: str) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for .xlsx extraction") from exc

    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        table_paths = []
        summaries = [
            f"source_name: {source_path.name}",
            "source_type: xlsx",
        ]
        for worksheet in workbook.worksheets:
            sheet_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", worksheet.title).strip("._")
            table_path = dirs["tables"] / f"{stem}.{sheet_name or 'sheet'}.csv"
            rows = list(worksheet.iter_rows(values_only=True))
            with table_path.open("w", encoding="utf-8", newline="") as dst:
                writer = csv.writer(dst)
                writer.writerows(
                    ["" if cell is None else cell for cell in row]
                    for row in rows
                )
            table_paths.append(str(table_path))
            summaries.append(
                f"sheet: {worksheet.title}, rows: {len(rows)}, columns: {worksheet.max_column}"
            )

        text_path = dirs["text"] / f"{stem}.txt"
        write_text(text_path, "\n".join(summaries) + "\n")
        return {"text": str(text_path), "tables": table_paths}
    except Exception as exc:
        # 손상된 XML 등으로 openpyxl이 workbook 자체를 열지 못하는 파일은
        # 매 배치마다 영구히 재시도되는 것을 막기 위해 기본 메타데이터로 대체한다.
        print(f"xlsx 파싱 실패, 기본 메타데이터로 대체: {source_path.name}: {exc}", file=sys.stderr)
        outputs = extract_basic_metadata(source_path, dirs, stem)
        text_path = Path(outputs["text"])
        note = f"content_extraction_error: {exc}\n"
        with text_path.open("a", encoding="utf-8") as f:
            f.write(note)
        return outputs


def extract_image(source_path: Path, dirs: dict[str, Path], stem: str) -> dict:
    image_info = {
        "source_name": source_path.name,
        "source_type": source_path.suffix.lower().lstrip("."),
        "size_bytes": source_path.stat().st_size,
        "path": str(source_path),
    }
    try:
        from PIL import Image
        from PIL.ExifTags import TAGS, GPSTAGS

        with Image.open(source_path) as image:
            image_info.update({
                "width": image.width,
                "height": image.height,
                "mode": image.mode,
                "format": image.format,
            })
            # EXIF 추출
            raw_exif = image._getexif() if hasattr(image, "_getexif") else None
            if raw_exif:
                exif = {TAGS.get(k, k): v for k, v in raw_exif.items()}
                if "DateTimeOriginal" in exif:
                    image_info["datetime_taken"] = str(exif["DateTimeOriginal"])
                if "Make" in exif:
                    image_info["camera_make"] = str(exif["Make"])
                if "Model" in exif:
                    image_info["camera_model"] = str(exif["Model"])
                # GPS 좌표 추출
                gps_info = exif.get("GPSInfo")
                if isinstance(gps_info, dict):
                    gps = {GPSTAGS.get(k, k): v for k, v in gps_info.items()}
                    def to_deg(val):
                        d, m, s = val
                        return float(d) + float(m) / 60 + float(s) / 3600
                    try:
                        lat = to_deg(gps["GPSLatitude"])
                        if gps.get("GPSLatitudeRef") == "S":
                            lat = -lat
                        lon = to_deg(gps["GPSLongitude"])
                        if gps.get("GPSLongitudeRef") == "W":
                            lon = -lon
                        image_info["gps_lat"] = round(lat, 6)
                        image_info["gps_lon"] = round(lon, 6)
                    except Exception:
                        pass
    except ImportError:
        image_info["note"] = "Pillow not installed; image dimensions not extracted."
    except Exception as exc:
        image_info["error"] = str(exc)

    table_path = dirs["tables"] / f"{stem}.image.json"
    table_path.write_text(
        json.dumps(image_info, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    text_path = dirs["text"] / f"{stem}.txt"
    lines = [f"{key}: {value}" for key, value in image_info.items()]
    write_text(text_path, "\n".join(lines) + "\n")
    return {"text": str(text_path), "tables": [str(table_path)]}

def extract_basic_metadata(source_path: Path, dirs: dict[str, Path], stem: str) -> dict:
    text_path = dirs["text"] / f"{stem}.txt"
    details = [
        f"source_name: {source_path.name}",
        f"source_type: {source_path.suffix.lower().lstrip('.') or 'unknown'}",
        f"size_bytes: {source_path.stat().st_size}",
        "content_extraction: unsupported_file_type",
    ]
    write_text(text_path, "\n".join(details) + "\n")
    return {"text": str(text_path), "tables": []}


def send_multipart_file(url: str, file_path: Path, extra_fields: dict | None = None) -> bytes:
    boundary = f"----WebKitFormBoundary{uuid.uuid4().hex}"
    parts = []

    if extra_fields:
        for name, value in extra_fields.items():
            parts.append(f"--{boundary}")
            parts.append(f'Content-Disposition: form-data; name="{name}"')
            parts.append('')
            parts.append(str(value))

    parts.append(f"--{boundary}")
    filename = file_path.name
    suffix = file_path.suffix.lower()
    content_type = "application/octet-stream"
    if suffix in {".jpg", ".jpeg"}:
        content_type = "image/jpeg"
    elif suffix == ".png":
        content_type = "image/png"
    elif suffix == ".pdf":
        content_type = "application/pdf"

    parts.append(f'Content-Disposition: form-data; name="file"; filename="{filename}"')
    parts.append(f'Content-Type: {content_type}')
    parts.append('')

    body_header = "\n".join(parts).replace("\n", "\r\n").encode("utf-8") + b"\r\n"
    body_footer = f"\r\n--{boundary}--\r\n".encode("utf-8")

    file_content = file_path.read_bytes()
    body = body_header + file_content + body_footer

    req = urllib.request.Request(
        url,
        data=body,
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Content-Length": str(len(body))
        },
        method="POST"
    )

    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def extract_via_ocr(source_path: Path, dirs: dict[str, Path], stem: str, ocr_endpoint: str, is_pdf: bool) -> dict:
    if is_pdf:
        url = f"{ocr_endpoint.rstrip('/')}/tools/pdf-ocr"
        extra = {"pages": "1"}
    else:
        url = f"{ocr_endpoint.rstrip('/')}/tools/image-ocr"
        extra = {}

    try:
        resp_bytes = send_multipart_file(url, source_path, extra)
        resp_json = json.loads(resp_bytes.decode("utf-8"))

        if resp_json.get("status") != "ok" or "url" not in resp_json:
            raise RuntimeError(f"OCR API failed: {resp_json}")

        file_url_path = resp_json["url"]
        download_url = f"{ocr_endpoint.rstrip('/')}{file_url_path}"

        excel_path = dirs["tables"] / f"{stem}_ocr.xlsx"
        with urllib.request.urlopen(download_url, timeout=30) as resp:
            excel_path.write_bytes(resp.read())

        from openpyxl import load_workbook
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        summaries = [
            f"source_name: {source_path.name}",
            f"source_type: {source_path.suffix.lower().lstrip('.')}",
            f"ocr_engine: gigabyte-atom-ocr",
            f"ocr_excel_source: {download_url}",
        ]

        for worksheet in workbook.worksheets:
            summaries.append(f"\n--- Sheet: {worksheet.title} ---")
            for row in worksheet.iter_rows(values_only=True):
                row_str = ", ".join(
                    ["" if cell is None else str(cell).strip() for cell in row]
                ).strip(", ")
                if row_str:
                    summaries.append(row_str)

        text_path = dirs["text"] / f"{stem}.txt"
        write_text(text_path, "\n".join(summaries) + "\n")

        return {"text": str(text_path), "tables": [str(excel_path)]}

    except Exception as exc:
        if not is_pdf:
            fallback_res = extract_image(source_path, dirs, stem)
            t_path = Path(fallback_res["text"])
            t_content = t_path.read_text(encoding="utf-8")
            write_text(t_path, t_content + f"ocr_error: {exc}\n")
            return fallback_res
        else:
            fallback_res = extract_basic_metadata(source_path, dirs, stem)
            t_path = Path(fallback_res["text"])
            t_content = t_path.read_text(encoding="utf-8")
            write_text(t_path, t_content + f"ocr_error: {exc}\n")
            return fallback_res


def extract_via_marker(source_path: Path, dirs: dict[str, Path], stem: str) -> dict:
    """Marker-PDF 엔진을 호출하여 PDF를 정밀 마크다운 텍스트로 추출합니다."""
    text_path = dirs["text"] / f"{stem}.txt"
    tmp_out_dir = dirs["text"] / f"tmp_{stem}"
    
    # 아톰 서버 venv 경로 또는 시스템 환경에 설치된 marker_single 바이너리 위치 추적
    marker_bin = "/home/caiser77/dgx_workspace/venv/bin/marker_single"
    if not Path(marker_bin).exists():
        import shutil
        marker_bin = shutil.which("marker_single") or "marker_single"

    import subprocess
    cmd = [
        marker_bin,
        str(source_path),
        "--output_dir",
        str(tmp_out_dir),
        "--output_format",
        "markdown"
    ]
    
    try:
        # 최초 구동 시 모델 다운로드 등으로 시간이 걸릴 수 있으므로 600초 타임아웃
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=600, encoding="utf-8")
        if result.returncode != 0:
            raise RuntimeError(f"Marker CLI execution failed: {result.stderr}")
            
        # 변환 완료된 md 파일 찾기 (marker_single은 output_dir/폴더명/파일명.md 형태로 저장함)
        expected_md_file = tmp_out_dir / source_path.stem / f"{source_path.stem}.md"
        if not expected_md_file.exists():
            expected_md_file = tmp_out_dir / stem / f"{stem}.md"
            
        if not expected_md_file.exists():
            md_files = list(tmp_out_dir.glob("**/*.md"))
            if md_files:
                expected_md_file = md_files[0]
                
        if expected_md_file.exists():
            md_content = expected_md_file.read_text(encoding="utf-8")
            write_text(text_path, md_content)
            
            import shutil
            shutil.rmtree(tmp_out_dir, ignore_errors=True)
            return {"text": str(text_path), "tables": []}
        else:
            raise FileNotFoundError("Parsed markdown output not found from Marker")
            
    except Exception as exc:
        print(f"Marker extraction failed, falling back to basic extraction: {exc}", file=sys.stderr)
        return extract_basic_metadata(source_path, dirs, stem)


def run_extraction(source_path: Path, output_dir: Path, no_ocr: bool = False, ocr_endpoint: str | None = None) -> dict:
    dirs = ensure_dirs(output_dir)
    stem = safe_stem(source_path)
    suffix = source_path.suffix.lower()

    if suffix in TEXT_EXTENSIONS:
        outputs = extract_text(source_path, dirs, stem)
    elif suffix in CSV_EXTENSIONS:
        outputs = extract_csv(source_path, dirs, stem)
    elif suffix in XLSX_EXTENSIONS:
        outputs = extract_xlsx(source_path, dirs, stem)
    elif suffix in IMAGE_EXTENSIONS:
        outputs = extract_image(source_path, dirs, stem)
    elif suffix in PDF_EXTENSIONS:
        # 로컬 아톰 서버에서는 고성능 Marker로 파싱을 우선 시도
        outputs = extract_via_marker(source_path, dirs, stem)
    else:
        outputs = extract_basic_metadata(source_path, dirs, stem)

    metadata_path = dirs["metadata"] / f"{stem}.metadata.json"
    outputs["metadata"] = str(metadata_path)
    return outputs


def build_metadata(
    source_path: Path,
    device_name: str,
    detected_at: str | None,
    status: str,
    outputs: dict,
    error: dict | None,
) -> dict:
    return {
        "source_path": str(source_path),
        "source_name": source_path.name,
        "source_type": source_path.suffix.lower().lstrip(".") or "unknown",
        "device_name": device_name,
        "detected_at": detected_at or now_iso(),
        "processed_at": now_iso(),
        "status": status,
        "outputs": outputs,
        "error": error,
    }


def write_metadata(metadata: dict, output_dir: Path, source_path: Path) -> Path:
    dirs = ensure_dirs(output_dir)
    output_metadata = metadata.get("outputs", {}).get("metadata")
    metadata_path = Path(
        output_metadata or dirs["metadata"] / f"{safe_stem(source_path)}.metadata.json"
    )
    metadata_path.parent.mkdir(parents=True, exist_ok=True)
    metadata_path.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return metadata_path


import sqlite3

def query_user_rules(file_path: Path) -> str:
    """SQLite DB에서 해당 파일 경로에 대한 기존 사용자 교정 규칙/피드백이 있는지 FTS5 또는 LIKE 검색을 수행합니다."""
    db_path = Path("/home/caiser77/dgx_workspace/004. 에르메스 NAS 분류기/data/aurum_nas_rules.db")
    if not db_path.exists():
        return ""

    rules_text = []
    try:
        conn = sqlite3.connect(str(db_path))
        cursor = conn.cursor()

        parts = file_path.parent.parts
        search_terms = []
        for i in range(1, min(len(parts), 4)):
            search_terms.append("/".join(parts[-i:]))

        for term in search_terms:
            cursor.execute(
                """
                SELECT user_approved_class, user_instruction
                FROM aurum_nas_rules
                WHERE original_path LIKE ?
                  AND user_approved_class IS NOT NULL
                  AND user_approved_class != ''
                ORDER BY feedback_at DESC, id DESC
                LIMIT 5
                """,
                (f"%{term}%",)
            )
            rows = cursor.fetchall()
            for row in rows:
                approved, inst = row
                rule = f"- 해당 경로 계층('{term}')의 이전 분류 정정: AI가 분류하려 했던 것 대신 사용자는 '{approved}'(으)로 승인했습니다."
                if inst:
                    rule += f" (사용자 특별 지시: {inst})"
                if rule not in rules_text:
                    rules_text.append(rule)

        conn.close()
    except Exception as e:
        print(f"Failed to query user rules from SQLite: {e}", file=sys.stderr)

    if rules_text:
        return "\n[과거 사용자 교정 피드백 / 학습 규칙]\n" + "\n".join(rules_text) + "\n"
    return ""


def build_feedback_preview(extracted_text: str, max_chars: int = 700) -> str:
    """텔레그램 승인 요청과 월요일 대기 큐에 넣을 문서 판단 근거 미리보기를 만듭니다."""
    if not extracted_text:
        return "본문 미리보기를 추출하지 못했습니다."
    lines = [line.strip() for line in extracted_text.splitlines() if line.strip()]
    preview = "\n".join(lines[:12]).strip()
    if len(preview) > max_chars:
        preview = preview[:max_chars].rstrip() + "..."
    return preview or "본문 미리보기를 추출하지 못했습니다."


def build_pending_feedback_payload(file_path: Path, ai_meta: dict, extracted_text: str = "") -> dict:
    project_name = ai_meta.get("project_name") or "미정"
    class_name = ai_meta.get("class_name") or "미정"
    question = ai_meta.get("question") or (
        f"'{file_path.name}' 파일의 사업명/분류군을 확정해야 합니다. "
        f"현재 추론은 사업명 '{project_name}', 분류군 '{class_name}'입니다. "
        "맞으면 승인, 틀리면 올바른 사업명과 분류군을 답장으로 알려주세요."
    )
    return {
        "created_at": now_iso(),
        "source_name": file_path.name,
        "original_path": str(file_path),
        "year_vendor": ai_meta.get("year_vendor") or "미정",
        "project_name": project_name,
        "inferred_class": class_name,
        "tags": ai_meta.get("tags") or [],
        "summary": ai_meta.get("summary") or "AI 요약이 비어 있습니다.",
        "question": question,
        "preview": build_feedback_preview(extracted_text),
        "status": "pending",
    }


def record_pending_feedback(file_path: Path, ai_meta: dict, extracted_text: str = "") -> dict:
    """월요일 브리핑용 분류 대기 큐를 누적 저장합니다."""
    pending_data = build_pending_feedback_payload(file_path, ai_meta, extracted_text)
    pending_dir = Path("/home/caiser77/dgx_workspace/004. 에르메스 NAS 분류기/data")
    pending_dir.mkdir(parents=True, exist_ok=True)
    (pending_dir / "last_pending.json").write_text(
        json.dumps(pending_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    with (pending_dir / "pending_queue.jsonl").open("a", encoding="utf-8") as pf:
        pf.write(json.dumps(pending_data, ensure_ascii=False) + "\n")
    return pending_data


def send_telegram_feedback_request(file_path: Path, ai_meta: dict, extracted_text: str = "") -> None:
    """AI 분류가 모호하거나 추가 확인이 필요한 경우, 텔레그램 봇을 통해 승인 요청 메시지를 발송합니다."""
    pending_data = record_pending_feedback(file_path, ai_meta, extracted_text)
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("ALLOWED_USER_ID")
    if not token or not chat_id:
        print("Telegram feedback skipped: TELEGRAM_BOT_TOKEN or ALLOWED_USER_ID is not set.", file=sys.stderr)
        return

    msg = f"""⚠️ [분류 모호성 감지 - 사용자 피드백 요청]
* 파일명: {pending_data['source_name']}
* NAS 경로: {pending_data['original_path']}
* AI 추론 결과:
  - 년도_업체: {pending_data['year_vendor']}
  - 사업명: {pending_data['project_name']}
  - 분류군: {pending_data['inferred_class']}
  - 태그: {', '.join(pending_data['tags']) if pending_data['tags'] else '없음'}
* AI 요약: {pending_data['summary']}
* AI 질문/사유: {pending_data['question']}

* 판단 근거 미리보기:
{pending_data['preview']}

위 분류를 승인하시겠습니까?
정정 또는 예외 사항이 있다면 답장으로 지시어를 보내주세요.
(예: '조류로 분류하고, 충청내륙 사업이야')
"""
    import urllib.request
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    data = {
        "chat_id": chat_id,
        "text": msg
    }

    try:
        req = urllib.request.Request(
            url,
            data=json.dumps(data).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            resp.read()
            print("Telegram feedback request sent successfully.")
    except Exception as e:
        print(f"Failed to send Telegram request: {e}", file=sys.stderr)


def clean_path_label(value: str, strip_numbering: bool = True) -> str:
    text = (value or "").replace("+", " ").strip()
    if strip_numbering:
        text = re.sub(r"^\d+(?:\.\d+)*[._\s-]+", "", text).strip()
    return text


def infer_structured_labels(file_path: Path, watch_path: str = "/mnt/dgxbackup") -> dict:
    """경로와 파일명에서 안정적인 구조 필드를 분리 추정합니다."""
    try:
        rel = file_path.relative_to(watch_path)
        parts = rel.parts
    except ValueError:
        raw_parts = file_path.parts
        if len(raw_parts) >= 4 and raw_parts[1] == "mnt":
            parts = raw_parts[3:]
        else:
            parts = raw_parts[1:] if raw_parts and raw_parts[0] == "/" else raw_parts

    if len(parts) >= 3 and re.fullmatch(r"20\d{2}", clean_path_label(parts[0], strip_numbering=False)):
        year = clean_path_label(parts[0], strip_numbering=False)
        vendor = clean_path_label(parts[1], strip_numbering=False)
        year_vendor = vendor if vendor.startswith(year) else f"{year} {vendor}".strip()
        project_name = clean_path_label(parts[2], strip_numbering=False)
    else:
        year_vendor = clean_path_label(parts[0], strip_numbering=False) if len(parts) > 0 else "unknown"
        project_name = clean_path_label(parts[1], strip_numbering=False) if len(parts) > 1 else "unknown"

    cleaned_parts = [clean_path_label(part) for part in parts]
    searchable = " ".join(cleaned_parts)

    document_type = "미확정"
    for keyword in DOCUMENT_TYPE_KEYWORDS:
        if keyword in searchable:
            document_type = keyword
            break

    taxon_group = "미확정"
    taxon_searchable = searchable.replace("동식물상", " ")
    for group in TAXON_GROUPS:
        if group in taxon_searchable:
            taxon_group = group
            break

    if document_type == "미확정":
        ext = file_path.suffix.lower().lstrip(".") or "unknown"
        document_type = ext.upper()

    tags = [f"#{year_vendor}", f"#{project_name}"]
    if taxon_group != "미확정":
        tags.append(f"#{taxon_group}")
    if document_type != "미확정":
        clean_doc_tag = re.sub(r"[^A-Za-z0-9가-힣_]+", "", document_type).strip()
        if clean_doc_tag:
            tags.append(f"#{clean_doc_tag}")

    return {
        "year_vendor": year_vendor,
        "project_name": project_name,
        "taxon_group": taxon_group,
        "document_type": document_type,
        "class_name": taxon_group if taxon_group != "미확정" else document_type,
        "tags": tags,
        "summary": f"{project_name} 관련 {document_type} 파일",
        "is_ambiguous": taxon_group == "미확정",
        "question": "본문 또는 경로에서 생물 분류군을 확정하지 못했습니다." if taxon_group == "미확정" else "",
    }


def normalize_classification(ai_meta: dict, file_path: Path) -> dict:
    """기존 class_name 중심 결과를 taxon_group/document_type 분리 구조로 정규화합니다."""
    base = infer_structured_labels(file_path)
    if not ai_meta or not isinstance(ai_meta, dict):
        return base

    merged = {**base, **ai_meta}
    class_name = str(merged.get("class_name") or "").strip()

    if not merged.get("taxon_group"):
        merged["taxon_group"] = class_name if class_name in TAXON_GROUPS else base["taxon_group"]
    if not merged.get("document_type"):
        merged["document_type"] = class_name if class_name in DOCUMENT_TYPE_KEYWORDS else base["document_type"]

    if merged.get("taxon_group") in DOCUMENT_TYPE_KEYWORDS:
        merged["document_type"] = merged["taxon_group"]
        merged["taxon_group"] = base["taxon_group"]

    if class_name in DOCUMENT_TYPE_KEYWORDS and base["taxon_group"] == "미확정":
        merged["document_type"] = class_name
        merged["taxon_group"] = "미확정"

    merged_taxon = merged.get("taxon_group")
    merged_doc = merged.get("document_type")
    merged["class_name"] = merged_taxon if merged_taxon and merged_taxon != "미확정" else (merged_doc or class_name)

    tags = []
    for tag in base.get("tags", []) + merged.get("tags", []):
        if tag and tag not in tags:
            tags.append(tag)
    merged["tags"] = tags
    return merged


def classify_from_path(file_path: Path, watch_path: str = "/mnt/dgxbackup") -> dict:
    """경로 구조에서 메타데이터 추출 — AI 없이 즉시 분류."""
    return infer_structured_labels(file_path, watch_path)


def has_usable_extracted_text(extracted_text: str) -> bool:
    """본문 추출 실패 안내문만 있는 경우 AI 분류 근거로 쓰지 않습니다."""
    if not extracted_text or not extracted_text.strip():
        return False
    failure_markers = (
        "content_extraction: unsupported_file_type",
        "ocr_error:",
    )
    if not any(marker in extracted_text for marker in failure_markers):
        return True

    ignored_prefixes = ("source_name:", "source_type:", "size_bytes:", "content_extraction:", "ocr_error:")
    meaningful_lines = [
        line.strip()
        for line in extracted_text.splitlines()
        if line.strip() and not line.strip().startswith(ignored_prefixes)
    ]
    return bool(meaningful_lines)


def classify_via_hermes(file_path: Path, extracted_text: str) -> dict:
    """Hermes 에이전트를 원샷 호출하여 NAS 데이터 지능형 분류를 수행합니다."""
    truncated_text = extracted_text[:2500] if extracted_text else ""
    if not has_usable_extracted_text(truncated_text):
        return classify_from_path(file_path)

    user_rules = query_user_rules(file_path)

    # 프롬프트 조립 (JSON 형식을 엄격하게 강제 및 최우선 순위 지침 삽입)
    prompt = f"""[중요 지침 - 최우선 순위]
아래에 제공되는 [과거 사용자 교정 피드백 / 학습 규칙]은 사용자가 이미 직접 검증하고 수정한 내용입니다.
과거 학습 규칙이 존재한다면, 본문 내용이나 경로명이 상충되더라도 반드시 사용자의 과거 피드백 지시사항을 100% 신뢰하여 분류 결과(year_vendor, project_name, taxon_group, document_type 등)를 결정해야 합니다.
그럴 경우 "is_ambiguous"는 무조건 false로 설정하고 "question"은 빈 문자열("")로 리턴하십시오. 질문을 던지거나 모호성을 true로 중복 보고해서는 안 됩니다.

{user_rules}

위 최우선 지침과 aurum_nas_classifier 스킬 규격에 맞추어 다음 입력 데이터를 분석하고 JSON으로만 응답해 주세요.
반드시 마크다운 코드 블록(```json ...) 없이 순수 JSON 텍스트만 출력해야 합니다.

분류 원칙:
- taxon_group은 생물 분류군만 입력합니다: 양서파충류, 조류, 포유류, 어류, 식물상, 육상곤충, 저서성대형무척추동물, 담수조류, 해충, 미확정.
- document_type은 문서/자료 유형만 입력합니다: 동식물상, 부록, 사업개요, 검토의견, 문헌, 현장사진, 사진, 야장, 영수증, 보고서, 조사표 등.
- 동식물상, 부록, 사업의 개요, 검토의견은 생물 분류군이 아니므로 taxon_group에 넣지 말고 document_type에 넣습니다.
- 생물 분류군 근거가 없으면 taxon_group은 반드시 "미확정"으로 둡니다.

입력 데이터:
{{
  "file_path": "{file_path}",
  "extracted_text": {json.dumps(truncated_text, ensure_ascii=False)}
}}

필수 출력 JSON 포맷 규격:
{{
  "year_vendor": "추론된 년도 및 업체 (예: 2026 LH)",
  "project_name": "추론된 사업명 (예: 인천용역)",
  "taxon_group": "생물 분류군 또는 미확정",
  "document_type": "문서/자료 유형",
  "class_name": "하위 호환용 대표 분류값: taxon_group이 확정이면 taxon_group, 아니면 document_type",
  "tags": ["#태그1", "#태그2"],
  "summary": "본문 내용 요약 (3줄 이내)",
  "is_ambiguous": false,
  "question": "모호할 경우 사용자에 보낼 질문 (비어있음 가능)"
}}"""

    hermes_bin = "/home/caiser77/hermes-agent/.venv/bin/hermes"
    if not Path(hermes_bin).exists():
        # 로컬(개발 환경) 테스트 대응용 Mock
        return {
            "year_vendor": "unknown",
            "project_name": "unknown",
            "class_name": "unknown",
            "tags": [],
            "summary": "Hermes binary not found locally.",
            "is_ambiguous": False,
            "question": ""
        }

    import subprocess
    cmd = [
        hermes_bin,
        "--oneshot",
        prompt,
        "--skills",
        "aurum_nas_classifier"
    ]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=90,
            encoding="utf-8"
        )
        if result.returncode != 0:
            print(f"Hermes classification failed with exit code {result.returncode}: {result.stderr}", file=sys.stderr)
            return {}

        stdout_clean = result.stdout.strip()
        # { 와 } 사이의 가장 긴 JSON 형태를 추출하여 파싱 시도 (주석 및 사족 방어)
        match = re.search(r'(\{.*\}).*', stdout_clean, re.DOTALL)
        if match:
            json_str = match.group(1)
            try:
                return normalize_classification(json.loads(json_str), file_path)
            except json.JSONDecodeError:
                # 개행 문자 등을 이스케이프하여 2차 파싱 시도
                try:
                    return normalize_classification(json.loads(json_str.replace('\n', '\\n')), file_path)
                except Exception:
                    pass
        return normalize_classification(json.loads(stdout_clean), file_path)
    except Exception as e:
        print(f"Error during Hermes classification: {e}", file=sys.stderr)
        print(f"Raw Hermes stdout was: {result.stdout if 'result' in locals() else 'None'}", file=sys.stderr)
        return {}


def generate_obsidian_note(metadata: dict, output_dir: Path) -> None:
    notes_dir = output_dir / "obsidian_notes"
    notes_dir.mkdir(parents=True, exist_ok=True)

    source_path = Path(metadata["source_path"])
    stem = safe_stem(source_path)
    note_path = notes_dir / f"{stem}.md"

    remote_watch_path = Path("/mnt/dgxbackup").resolve()
    mac_watch_path = "/Users/nams/AI_BASE/Obsi-26Project_2026"

    tags = ["#자동색인", "#NAS_데이터"]
    path_meta = {}
    rel_path = source_path

    try:
        rel_path = source_path.resolve().relative_to(remote_watch_path)
        mac_link_path = f"file://{mac_watch_path}/{rel_path}"

        # 경로 조각들(폴더 계층)을 파싱하여 태그화
        parts = rel_path.parts[:-1]
        for part in parts:
            clean_tag = re.sub(r"[^A-Za-z0-9가-힣_]+", "", part).strip()
            if clean_tag:
                tags.append(f"#{clean_tag}")

        if len(parts) >= 1:
            path_meta["년도_업체"] = parts[0].strip()
        if len(parts) >= 2:
            path_meta["사업명"] = parts[1].strip()
        if len(parts) >= 3:
            inferred = infer_structured_labels(source_path)
            path_meta["생물분류군"] = inferred.get("taxon_group", "미확정")
            path_meta["문서유형"] = inferred.get("document_type", parts[2].strip())
            path_meta["분류군"] = inferred.get("taxon_group", "미확정")
    except ValueError:
        mac_link_path = f"file://{source_path}"

    # AI 분류 결과가 메타데이터에 존재하는지 판별
    ai_meta = metadata.get("ai_classification", {})
    if ai_meta and isinstance(ai_meta, dict):
        ai_tags = ai_meta.get("tags", [])
        for t in ai_tags:
            t_clean = t.strip()
            if not t_clean.startswith("#"):
                t_clean = f"#{t_clean}"
            if t_clean not in tags:
                tags.append(t_clean)

        if ai_meta.get("year_vendor"):
            path_meta["년도_업체"] = ai_meta["year_vendor"]
        if ai_meta.get("project_name"):
            path_meta["사업명"] = ai_meta["project_name"]
        if ai_meta.get("taxon_group"):
            path_meta["생물분류군"] = ai_meta["taxon_group"]
            path_meta["분류군"] = ai_meta["taxon_group"]
        elif ai_meta.get("class_name"):
            path_meta["분류군"] = ai_meta["class_name"]
        if ai_meta.get("document_type"):
            path_meta["문서유형"] = ai_meta["document_type"]

    extracted_text_path = metadata.get("outputs", {}).get("text")
    preview = ""
    if extracted_text_path and Path(extracted_text_path).exists():
        try:
            content = Path(extracted_text_path).read_text(encoding="utf-8")
            preview = "\n".join(content.splitlines()[:15])
        except Exception:
            preview = "미리보기를 불러올 수 없습니다."

    # YAML 메타데이터 구성
    yaml_lines = [
        "---",
        f"생성일: {metadata['processed_at'][:10]}",
        f"장비명: {metadata['device_name']}",
        f"원본이름: {metadata['source_name']}",
        f"원본유형: {metadata['source_type']}"
    ]
    for key, val in path_meta.items():
        yaml_lines.append(f"{key}: {val}")

    if ai_meta:
        yaml_lines.append(f"AI_요약: {ai_meta.get('summary', '').replace(chr(10), ' ')}")
        yaml_lines.append(f"AI_모호성: {str(ai_meta.get('is_ambiguous', False)).lower()}")
        if ai_meta.get("question"):
            yaml_lines.append(f"AI_질문: {ai_meta.get('question', '').replace(chr(10), ' ')}")

    yaml_lines.append(f"태그: {' '.join(tags)}")
    yaml_lines.append("---")

    path_meta_content = ""
    if path_meta:
        path_meta_content = "\n".join(f"* **{k}**: {v}" for k, v in path_meta.items()) + "\n"

    yaml_content = "\n".join(yaml_lines)
    note_content = f"""{yaml_content}
# 📄 {metadata["source_name"]} 요약본

* **처리 시각**: {metadata["processed_at"]}
{path_meta_content}* **파일 상대경로**: {rel_path}
"""
    if ai_meta and ai_meta.get("summary"):
        note_content += f"""
---
### 🤖 AI 분석 요약
> {ai_meta.get("summary")}
"""

    note_content += f"""
---
### 🔍 문서 미리보기 (첫 15줄)

```text
{preview}
```

---
### 🔗 NAS 원본 파일 링크
* 🖥️ 맥북 로컬 열기: [{metadata["source_name"]} 열기]({mac_link_path})
"""
    note_path.write_text(note_content, encoding="utf-8")


def process_image_directory(
    dir_path: Path,
    output_dir: Path,
    device_name: str,
    detected_at: str | None = None,
    num_threads: int = 16,
) -> int:
    """디렉토리 내의 모든 사진 파일을 찾아 EXIF 정보를 병렬로 고속 추출하고 통합 CSV와 Obsidian 노트를 생성합니다."""
    import concurrent.futures
    import os
    from PIL import Image
    from PIL.ExifTags import TAGS, GPSTAGS

    # 1. 대상 사진 파일 수집
    image_paths = []
    # os.walk를 활용해 recursive하게 이미지 수집
    for root, subdirs, files in os.walk(dir_path):
        # 숨김 디렉토리 제외
        subdirs[:] = [d for d in subdirs if not d.startswith('.')]
        for f in files:
            if f.startswith('.'):
                continue
            path = Path(root) / f
            if path.suffix.lower() in IMAGE_EXTENSIONS:
                image_paths.append(path)

    total_images = len(image_paths)
    if total_images == 0:
        print(f"디렉토리에서 이미지 파일을 찾을 수 없습니다: {dir_path}")
        return 0

    print(f"총 {total_images}개의 이미지 파일 감지. {num_threads}개 스레드로 고속 병렬 EXIF 추출을 개시합니다...")

    dirs = ensure_dirs(output_dir)
    results = []
    success_count = 0
    failed_count = 0

    def extract_single(source_path: Path) -> dict | None:
        nonlocal success_count, failed_count
        stem = safe_stem(source_path)
        image_info = {
            "source_name": source_path.name,
            "source_type": source_path.suffix.lower().lstrip("."),
            "size_bytes": source_path.stat().st_size,
            "path": str(source_path),
            "width": None,
            "height": None,
            "mode": None,
            "format": None,
            "datetime_taken": "",
            "camera_make": "",
            "camera_model": "",
            "gps_lat": "",
            "gps_lon": "",
            "status": "success",
            "error": "",
        }
        try:
            with Image.open(source_path) as image:
                image_info.update({
                    "width": image.width,
                    "height": image.height,
                    "mode": image.mode,
                    "format": image.format,
                })
                raw_exif = image._getexif() if hasattr(image, "_getexif") else None
                if raw_exif:
                    exif = {TAGS.get(k, k): v for k, v in raw_exif.items()}
                    if "DateTimeOriginal" in exif:
                        image_info["datetime_taken"] = str(exif["DateTimeOriginal"])
                    if "Make" in exif:
                        image_info["camera_make"] = str(exif["Make"])
                    if "Model" in exif:
                        image_info["camera_model"] = str(exif["Model"])

                    # GPS 추출
                    gps_info = exif.get("GPSInfo")
                    if isinstance(gps_info, dict):
                        gps = {GPSTAGS.get(k, k): v for k, v in gps_info.items()}
                        def to_deg(val):
                            d, m, s = val
                            return float(d) + float(m) / 60 + float(s) / 3600
                        try:
                            lat = to_deg(gps["GPSLatitude"])
                            if gps.get("GPSLatitudeRef") == "S":
                                lat = -lat
                            lon = to_deg(gps["GPSLongitude"])
                            if gps.get("GPSLongitudeRef") == "W":
                                lon = -lon
                            image_info["gps_lat"] = round(lat, 6)
                            image_info["gps_lon"] = round(lon, 6)
                        except Exception:
                            pass
            success_count += 1
            return image_info
        except Exception as exc:
            image_info["status"] = "failed"
            image_info["error"] = str(exc)
            failed_count += 1
            return image_info

    # ThreadPoolExecutor를 이용해 병렬 처리
    processed_count = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
        future_to_path = {executor.submit(extract_single, path): path for path in image_paths}
        for future in concurrent.futures.as_completed(future_to_path):
            res = future.result()
            if res:
                results.append(res)
                processed_count += 1
                if processed_count % 100 == 0 or processed_count == total_images:
                    print(f"진행 상황: {processed_count}/{total_images} 완료...")

    # 2. 통합 CSV 파일 작성
    csv_file_path = output_dir / "image_metadata_summary.csv"
    headers = [
        "source_name", "source_type", "size_bytes", "path", "width", "height",
        "mode", "format", "datetime_taken", "camera_make", "camera_model",
        "gps_lat", "gps_lon", "status", "error"
    ]

    try:
        with csv_file_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for res in results:
                writer.writerow(res)
        print(f"통합 CSV 메타데이터 저장 완료: {csv_file_path}")
    except Exception as e:
        print(f"CSV 저장 중 오류 발생: {e}", file=sys.stderr)

    # 3. 각 이미지 메타데이터 파일 및 Obsidian 노트 생성
    print("각 이미지의 Obsidian 노트를 생성하고 있습니다...")
    obsidian_count = 0
    for res in results:
        if res["status"] != "success":
            continue

        path_obj = Path(res["path"])
        stem = safe_stem(path_obj)

        meta = build_metadata(
            path_obj,
            device_name,
            detected_at,
            "success",
            {
                "text": str(dirs["text"] / f"{stem}.txt"),
                "tables": [str(dirs["tables"] / f"{stem}.image.json")]
            },
            None
        )
        meta["image_info"] = {
            "width": res["width"],
            "height": res["height"],
            "mode": res["mode"],
            "format": res["format"],
            "datetime_taken": res["datetime_taken"],
            "camera_make": res["camera_make"],
            "camera_model": res["camera_model"],
            "gps_lat": res["gps_lat"],
            "gps_lon": res["gps_lon"]
        }

        # 파일명 기반 생물종 분류 추정
        species_candidate = ""
        hangul_words = re.findall(r'[가-힣]{2,}', path_obj.name)
        if hangul_words:
            species_candidate = hangul_words[0]

        ai_meta = {
            "year_vendor": "2026 LH",
            "project_name": dir_path.name,
            "class_name": species_candidate if species_candidate else "사진",
            "tags": [f"#{species_candidate}"] if species_candidate else [],
            "summary": f"{species_candidate if species_candidate else '사진'} 이미지 파일. 파일명: {path_obj.name}",
            "is_ambiguous": False,
            "question": ""
        }

        if species_candidate:
            ai_meta["tags"].append("#특이종")
            ai_meta["tags"].append(f"#{species_candidate}")

        meta["ai_classification"] = ai_meta

        try:
            write_metadata(meta, output_dir, path_obj)

            text_lines = [f"{k}: {v}" for k, v in res.items()]
            write_text(dirs["text"] / f"{stem}.txt", "\n".join(text_lines) + "\n")

            (dirs["tables"] / f"{stem}.image.json").write_text(
                json.dumps(res, ensure_ascii=False, indent=2), encoding="utf-8"
            )

            generate_obsidian_note(meta, output_dir)
            obsidian_count += 1
        except Exception:
            pass

    print(f"Obsidian 노트 {obsidian_count}개 생성 완료.")
    print(f"=== 고속 배치 이미지 추출 완료: 성공 {success_count}개, 실패 {failed_count}개 ===")
    return 0



def export_to_collaboration_pipeline(metadata_path: Path, output_dir: Path, pipeline_root: str | None) -> None:
    """성공 metadata를 005 협업 파이프라인 raw_analyzed 입력으로 변환한다."""
    if not pipeline_root:
        return
    project_root = Path(__file__).resolve().parent.parent.parent
    exporter = project_root / "005. 아톰-모하비-아우룸 협업 파이프라인" / "scripts" / "export_processed_to_pipeline.py"
    if not exporter.exists():
        print(f"005 pipeline exporter not found: {exporter}", file=sys.stderr)
        return
    cmd = [
        sys.executable or "python3",
        str(exporter),
        "--processed-dir", str(output_dir),
        "--pipeline-root", str(Path(pipeline_root).expanduser()),
        "--metadata-file", str(metadata_path),
        "--overwrite",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.stdout.strip():
        print(result.stdout.strip())
    if result.returncode != 0:
        print(f"005 pipeline export failed: {result.stderr.strip()}", file=sys.stderr)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("legacy_input", nargs="?", help="이전 형식의 입력 파일 경로")
    parser.add_argument("--input", dest="input_path", help="처리할 입력 파일 경로")
    parser.add_argument("--device-name", default="unknown", help="장비명")
    parser.add_argument("--output-dir", default="data/processed", help="추출 결과 저장 루트")
    parser.add_argument("--detected-at", help="감지 시각 ISO 문자열")
    parser.add_argument("--ocr-endpoint", default="http://192.168.219.100:7870", help="기가바이트 OCR API 엔드포인트")
    parser.add_argument("--no-ocr", action="store_true", help="원격 OCR 연동을 생략하고 로컬 추출만 수행")
    parser.add_argument("--no-telegram", action="store_true", help="텔레그램 알림 발송 생략 (배치 모드용)")
    parser.add_argument("--no-ai", action="store_true", help="Hermes AI 분류 생략 — 경로 파싱으로 대체 (배치 고속 모드)")
    parser.add_argument("--threads", type=int, default=16, help="디렉토리 스캔 시 병렬 작업자(스레드) 수")
    parser.add_argument("--pipeline-root", help="005 협업 파이프라인 stage 루트. 지정 시 raw_analyzed 입력 파일을 생성")
    args = parser.parse_args()

    input_value = args.input_path or args.legacy_input
    if not input_value:
        parser.error("--input is required")

    source_path = Path(input_value).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser()
    if not output_dir.is_absolute():
        output_dir = Path.cwd() / output_dir

    if not source_path.exists():
        error = {
            "type": "FileNotFoundError",
            "message": f"input file not found: {source_path}",
            "return_code": 1,
        }
        metadata = build_metadata(
            source_path,
            args.device_name,
            args.detected_at,
            "failed",
            {"text": None, "tables": [], "metadata": None},
            error,
        )
        write_metadata(metadata, output_dir, source_path)
        print(error["message"], file=sys.stderr)
        return 1

    if source_path.is_dir():
        print(f"디렉토리 감지: {source_path}")
        print("이미지 고속 병렬 추출 프로세스를 실행합니다.")
        try:
            return process_image_directory(
                source_path,
                output_dir,
                args.device_name,
                args.detected_at,
                args.threads,
            )
        except Exception as e:
            print(f"디렉토리 이미지 추출 중 오류 발생: {e}", file=sys.stderr)
            return 1

    try:
        outputs = run_extraction(source_path, output_dir, args.no_ocr, args.ocr_endpoint)

        # text 본문 가져와서 Hermes AI 분류 수행
        text_output_path = outputs.get("text")
        extracted_text = ""
        if text_output_path and Path(text_output_path).exists():
            try:
                extracted_text = Path(text_output_path).read_text(encoding="utf-8")
            except Exception:
                pass

        ai_classification = {}
        no_ai = getattr(args, "no_ai", False)
        if no_ai:
            ai_classification = classify_from_path(source_path)
        elif not args.no_ocr:
            print("Running Hermes AI classification...")
            ai_classification = normalize_classification(classify_via_hermes(source_path, extracted_text), source_path)
            print(f"AI classification result: {ai_classification}")
            if ai_classification.get("is_ambiguous") and not getattr(args, "no_telegram", False):
                print("Ambiguity detected. Sending Telegram alert...")
                send_telegram_feedback_request(source_path, ai_classification, extracted_text)

        metadata = build_metadata(
            source_path,
            args.device_name,
            args.detected_at,
            "success",
            outputs,
            None,
        )
        metadata["ai_classification"] = ai_classification

        metadata_path = write_metadata(metadata, output_dir, source_path)
        try:
            generate_obsidian_note(metadata, output_dir)
        except Exception as e:
            print(f"Failed to generate obsidian note: {e}", file=sys.stderr)
        export_to_collaboration_pipeline(metadata_path, output_dir, args.pipeline_root)
        print(f"metadata={metadata_path}")
        return 0
    except Exception as exc:
        error = {
            "type": exc.__class__.__name__,
            "message": str(exc),
            "return_code": 1,
        }
        metadata = build_metadata(
            source_path,
            args.device_name,
            args.detected_at,
            "failed",
            {"text": None, "tables": [], "metadata": None},
            error,
        )
        metadata_path = write_metadata(metadata, output_dir, source_path)
        try:
            generate_obsidian_note(metadata, output_dir)
        except Exception as e:
            print(f"Failed to generate obsidian note: {e}", file=sys.stderr)
        print(f"extraction failed; metadata={metadata_path}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
