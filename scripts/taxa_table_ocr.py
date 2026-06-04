import sys
import re
from pathlib import Path
from difflib import SequenceMatcher

import cv2
import numpy as np
import pandas as pd
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

BASE = Path("/home/caiser77/dgx_workspace")
SPECIES_MAP_PATH = BASE / "data" / "taxa_species_map.csv"

HEADERS = ["학명", "국명", "1차", "2차", "3차", "4차", "5차", "종합", "비고", "원문국명", "매칭점수"]


def clean_text(text):
    text = str(text or "").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()

    for w in ["EIASS", "ZAPA", "BASE", "환경영향평가", "정보지원시스템"]:
        text = text.replace(w, "")

    return text.strip(" \"'`|")


def normalize_korean(text):
    text = clean_text(text)
    return re.sub(r"[^가-힣A-Za-z0-9]", "", text)


def load_species_map():
    df = pd.read_csv(SPECIES_MAP_PATH, encoding="utf-8-sig")

    records = []
    exact = {}

    for _, row in df.iterrows():
        sci = str(row["scientific_name"]).strip()
        kor = str(row["korean_name"]).strip()

        if not sci or not kor or sci == "nan" or kor == "nan":
            continue

        key = normalize_korean(kor)
        if not key:
            continue

        records.append((sci, kor, key))
        exact.setdefault(key, (sci, kor))

    return records, exact


def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()


def match_korean_name(raw, records, exact):
    raw_clean = clean_text(raw)
    raw_norm = normalize_korean(raw_clean)

    if not raw_norm:
        return "", "", raw_clean, 0.0

    if raw_norm in exact:
        sci, kor = exact[raw_norm]
        return sci, kor, raw_clean, 1.0

    best_sci = ""
    best_kor = ""
    best_score = 0.0

    for sci, kor, kor_norm in records:
        score = similarity(raw_norm, kor_norm)

        if raw_norm in kor_norm or kor_norm in raw_norm:
            score += 0.12

        if score > best_score:
            best_sci = sci
            best_kor = kor
            best_score = score

    if best_score >= 0.58:
        return best_sci, best_kor, raw_clean, min(best_score, 1.0)

    return "", "", raw_clean, best_score


def merge_positions(pos, gap=10):
    if len(pos) == 0:
        return []

    merged = []
    group = [int(pos[0])]

    for p in pos[1:]:
        p = int(p)
        if p - group[-1] <= gap:
            group.append(p)
        else:
            merged.append(int(np.mean(group)))
            group = [p]

    merged.append(int(np.mean(group)))
    return merged


def detect_grid(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    binary = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        25,
        12,
    )

    h, w = binary.shape

    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (max(w // 25, 25), 1)),
    )

    vertical = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(h // 35, 18))),
    )

    ys_raw = np.where(np.sum(horizontal > 0, axis=1) > w * 0.25)[0]
    xs_raw = np.where(np.sum(vertical > 0, axis=0) > h * 0.16)[0]

    xs = merge_positions(xs_raw, gap=10)
    ys = merge_positions(ys_raw, gap=10)

    return xs, ys


def crop_cell(image, xs, ys, r, c):
    x1, x2 = xs[c], xs[c + 1]
    y1, y2 = ys[r], ys[r + 1]

    pad = 2
    return image[y1 + pad:y2 - pad, x1 + pad:x2 - pad]


def preprocess_text(img, scale=3):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    gray[gray > 180] = 255
    gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)

    th = cv2.threshold(
        gray,
        0,
        255,
        cv2.THRESH_BINARY + cv2.THRESH_OTSU,
    )[1]

    return th


def ocr_korean(img):
    if img is None or img.size == 0:
        return ""

    th = preprocess_text(img, scale=3)

    results = []
    for psm in [7, 6]:
        txt = pytesseract.image_to_string(
            th,
            lang="kor",
            config=f"--psm {psm}",
        )
        txt = clean_text(txt)
        if txt:
            results.append(txt)

    if not results:
        return ""

    def score_result(x):
        hangul_count = len(re.findall(r"[가-힣]", x))
        return hangul_count - abs(len(x) - 5) * 0.2

    return max(results, key=score_result)


def ocr_scientific_light(img):
    if img is None or img.size == 0:
        return ""

    th = preprocess_text(img, scale=3)
    txt = pytesseract.image_to_string(th, lang="eng", config="--psm 7")
    return clean_text(txt)


def ocr_remark(img):
    if img is None or img.size == 0:
        return ""

    th = preprocess_text(img, scale=3)
    txt = pytesseract.image_to_string(th, lang="kor+eng", config="--psm 7")
    txt = clean_text(txt)

    if "교" in txt:
        return "교"

    if txt in ["0", "O", "o", "ㅇ"]:
        return ""

    return txt


def has_dot(img):
    if img is None or img.size == 0:
        return False

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray[gray > 200] = 255

    _, th = cv2.threshold(gray, 120, 255, cv2.THRESH_BINARY_INV)

    h, w = th.shape
    area_total = h * w

    contours, _ = cv2.findContours(th, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    for cnt in contours:
        area = cv2.contourArea(cnt)

        if area_total * 0.004 < area < area_total * 0.30:
            x, y, cw, ch = cv2.boundingRect(cnt)
            ratio = cw / max(ch, 1)

            cx = x + cw / 2
            cy = y + ch / 2

            if 0.45 <= ratio <= 2.2 and cw >= 3 and ch >= 3:
                if w * 0.12 <= cx <= w * 0.88 and h * 0.12 <= cy <= h * 0.88:
                    return True

    return False


def is_valid_korean_row(kor, raw, score):
    if not kor:
        return False

    if score < 0.58:
        return False

    raw_norm = normalize_korean(raw)

    if len(raw_norm) < 2:
        return False

    bad = ["학명", "국명", "비고", "출현종수", "목록", "조사지역", "현지조사"]
    if any(b in raw for b in bad):
        return False

    return True


def parse_table(image_path, output_xlsx):
    records, exact = load_species_map()

    image = cv2.imread(str(image_path))
    if image is None:
        raise FileNotFoundError(f"이미지를 읽을 수 없습니다: {image_path}")

    xs, ys = detect_grid(image)

    debug_path = Path("outputs") / f"{Path(image_path).stem}_grid_debug.txt"
    debug_path.write_text(
        f"xs={xs}\nys={ys}\ncols={len(xs)-1}\nrows={len(ys)-1}\n",
        encoding="utf-8",
    )

    if len(xs) < 10 or len(ys) < 5:
        raise RuntimeError(f"표 격자 감지 실패. debug: {debug_path}")

    rows = []

    for r in range(len(ys) - 1):
        row_cells = [crop_cell(image, xs, ys, r, c) for c in range(9)]

        raw_kor = ocr_korean(row_cells[1])
        sci, kor, raw_kor_clean, score = match_korean_name(raw_kor, records, exact)

        if not is_valid_korean_row(kor, raw_kor_clean, score):
            continue

        final_sci = sci if sci else ocr_scientific_light(row_cells[0])

        values = [final_sci, kor]

        for cell in row_cells[2:8]:
            values.append("●" if has_dot(cell) else "")

        values.append(ocr_remark(row_cells[8]))
        values.append(raw_kor_clean)
        values.append(round(float(score), 3))

        rows.append(values)

    df = pd.DataFrame(rows, columns=HEADERS)
    df.to_excel(output_xlsx, index=False)

    format_excel(output_xlsx)

    print(f"감지된 행 수: {len(rows)}")
    print(f"격자 디버그: {debug_path}")


def format_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "분류군 조사표 OCR"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws["A"][1:]:
        cell.font = Font(italic=True)

    widths = {
        "A": 45,
        "B": 22,
        "C": 8,
        "D": 8,
        "E": 8,
        "F": 8,
        "G": 8,
        "H": 8,
        "I": 10,
        "J": 20,
        "K": 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(path)


def main():
    if len(sys.argv) < 2:
        print("사용법: python scripts/taxa_table_ocr.py uploads/file.png")
        return

    inp = Path(sys.argv[1])
    out = Path("outputs") / f"{inp.stem}_taxa.xlsx"

    parse_table(inp, out)
    print(f"완료: {out}")


if __name__ == "__main__":
    main()
